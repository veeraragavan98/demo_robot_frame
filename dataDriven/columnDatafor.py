import openpyxl
from jproperties import Properties

#A1 = [1, 2, 3, 4]

A1= [1, 2, 3, 4]
B2= [3, 4, 5, 6, 7]

configs = Properties()
with open('C:/Users/BVeeraragavan/PycharmProjects/Acxhange/writeExcel.properties', 'rb') as config_file:
    configs.load(config_file)
    A = configs.get('A')
    B = configs.get('B')
#Load workbook
wb = openpyxl.load_workbook('C:/Users/BVeeraragavan/PycharmProjects/Acxhange/Excel/DataDriven.xlsx')
sheet = wb.active
#Itrate the row and inputs value
for i, Aw in enumerate(A, start=1):
    sheet.cell(row=1, column=i, value=Aw)
    print("*********")
#Itrate the row and inputs value
for j, item in enumerate(B, start=1):
    sheet.cell(row=2, column=j, value=item)
wb.save('C:/Users/BVeeraragavan/PycharmProjects/Acxhange/Excel/DataColumn12.xlsx')
